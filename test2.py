import openpyxl
import matplotlib.pyplot as plt

wb = openpyxl.load_workbook("test2.xlsx");
ws = wb['Sheet1'];

teacher = []
for tnum in range(2,16): #P열은 16열, 현 데이터는 O열까지 존재
    tvary = ws.cell(row=2, column=tnum).value
    teacher.append(tvary);
student = []
for snum in range(2,16):
    svary = ws.cell(row=3, column=snum).value
    student.append(svary);
ts = []
for tsnum in range(2,16):
    tsvary = ws.cell(row=4, column=tsnum).value
    ts.append(tsvary);
years = []
for year in range (2007,2021):
    years.append(year);
fig, ax1 = plt.subplots() #해석 필요
ax1.set_xlabel('Year', size='medium')
ax1.set_ylabel('the Number of Teacher', labelpad=10, color='yellowgreen', size='x-large')
ax1.plot(years, teacher, marker='o', color='yellowgreen', label='teacher') #교직원 수
ax2 = ax1.twinx() #y축 2개 (x축을 같이 사용한다)
ax2.set_ylabel('the Number of Student', labelpad=10, color='pink', size='x-large')
ax2.plot(years, student, marker='o', color='pink', label='student') #학생 수
#plt.plot(years, ts, marker='o', color='orange') #교원 1인 당 학생 
plt.title('The Change in the Number of Students and Teachers in Incheon')
plt.tight_layout(pad=0.4, w_pad=0.5, h_pad=1.0) #디자인적 요소
plt.grid() #격자점
plt.show()

