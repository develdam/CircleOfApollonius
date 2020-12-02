import openpyxl
import matplotlib.pyplot as plt

wb = openpyxl.load_workbook("test3.xlsx");
ws = wb['Sheet1'];

yi = int(input('TYPE the year : '))
if (yi < 1995 or yi > 2019):
    print("error");
    exit()
yr = yi - 1992

sun = ws.cell(column=yr, row=6).value
wind = ws.cell(column=yr, row=7).value
water = ws.cell(column=yr, row=8).value
ocean = ws.cell(column=yr, row=9).value
bio = ws.cell(column=yr, row=10).value
waste = ws.cell(column=yr, row=11).value
fuelcell = ws.cell(column=yr, row=12).value
igcc = ws.cell(column=yr, row=13).value
datas = [sun,wind,water,ocean,bio,waste,fuelcell,igcc]
categories = ['sun','wind','water','ocean','bio','waste','fuelcell','igcc']

plt.pie(datas,labels=categories)
plt.show()
